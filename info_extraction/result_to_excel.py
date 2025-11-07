#!/usr/bin/env python3
"""
Builds client-ready Excel summaries from previously scraped material data.

Expected folder layout (default BASE_DOWNLOAD_DIR):
downloads/
└── <source_excel_stem>/
    ├── material_<mano>/
    │   ├── material_<mano>_data.json
    │   └── images/
    │       ├── image_1.png
    │       └── ...
    └── material_<mano>/
        ...

Each <source_excel_stem> directory will receive a <source_excel_stem>_summary.xlsx file.
"""

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment

# --------------------------------------------------------------------------- #
# Configuration – update BASE_DOWNLOAD_DIR to match your environment
# --------------------------------------------------------------------------- #
BASE_DOWNLOAD_DIR = Path("/downloads")
SUMMARY_SUFFIX = "_summary.xlsx"


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def format_qa_requirements(qa_reqs: Optional[Dict[str, Any]]) -> str:
    """Return newline-delimited list of QA tests whose value is True."""
    if not qa_reqs:
        return ""
    true_tests = [
        name for name, value in qa_reqs.items()
        if isinstance(value, bool) and value
    ]
    if qa_reqs.get("Additional Tests", "") != "":
        true_tests.append("Additional Tests" + qa_reqs.get("Additional Tests", ""))
    if qa_reqs.get("Comments", "") != "":
        true_tests.append("Comments" + qa_reqs.get("Comments", ""))

    return "\n".join(true_tests)


def find_material_json(material_dir: Path) -> Optional[Path]:
    """Locate the JSON snapshot inside a material directory."""
    candidates = list(material_dir.glob("material_*_data.json"))
    return candidates[0] if candidates else None


def collect_material_entry(material_dir: Path) -> Optional[Dict[str, Any]]:
    """Parse the JSON + locate images for a single material."""
    json_path = find_material_json(material_dir)
    if not json_path:
        print(f"[WARN] JSON not found in {material_dir}")
        return None

    try:
        with open(json_path, "r") as f:
            data = json.load(f)
    except Exception as exc:
        print(f"[WARN] Failed to read {json_path}: {exc}")
        return None

    image_dir = material_dir / "images"
    image_paths = sorted(image_dir.glob("*")) if image_dir.exists() else []
    first_image = next((img for img in image_paths if img.is_file()), None)

    return {
        "material_number": data.get("material_number"),
        "component_id": data.get("component_id"),
        "cost": data.get("cost"),
        "supplier_name": data.get("supplier_name"),
        "supplier_material_no": data.get("supplier_material_no"),
        "qa_text": format_qa_requirements(data.get("qa_requirements")),
        "image_path": first_image,
    }


def write_summary_excel(entries: List[Dict[str, Any]], output_path: Path) -> None:
    """Write the collected entries to an Excel workbook with embedded images."""
    if not entries:
        print(f"[INFO] No material entries for {output_path.parent.name}; skipping Excel export.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Materials"

    headers = [
        "Material Number",
        "Component ID",
        "Cost",
        "Supplier Name",
        "Supplier Material NO",
        "QA Requirements (True)",
        "Image",
    ]
    ws.append(headers)

    column_widths = {
        "A": 18,
        "B": 15,
        "C": 10,
        "D": 40,
        "E": 20,
        "F": 60,
        "G": 20,
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row_idx, entry in enumerate(entries, start=2):
        ws.cell(row=row_idx, column=1, value=entry.get("material_number"))
        ws.cell(row=row_idx, column=2, value=entry.get("component_id"))
        ws.cell(row=row_idx, column=3, value=entry.get("cost"))
        ws.cell(row=row_idx, column=4, value=entry.get("supplier_name"))
        ws.cell(row=row_idx, column=5, value=entry.get("supplier_material_no"))

        qa_cell = ws.cell(row=row_idx, column=6, value=entry.get("qa_text") or "")
        qa_cell.alignment = Alignment(wrap_text=True, vertical="top")

        image_path = entry.get("image_path")
        if image_path and image_path.exists():
            try:
                img = XLImage(str(image_path))
                img.width = 120
                img.height = 120
                ws.add_image(img, f"G{row_idx}")
                ws.row_dimensions[row_idx].height = max(
                    ws.row_dimensions[row_idx].height or 0,
                    100,
                )
            except Exception as exc:
                print(f"[WARN] Failed to embed image {image_path}: {exc}")

    wb.save(output_path)
    print(f"[OK] Summary written to {output_path}")


# --------------------------------------------------------------------------- #
# Main workflow
# --------------------------------------------------------------------------- #
def generate_reports(base_download_dir: Path) -> None:
    """Build a summary workbook for each source Excel folder under base_download_dir."""
    if not base_download_dir.exists():
        raise FileNotFoundError(f"Base directory '{base_download_dir}' does not exist.")
    final_report_dir = base_download_dir / "summary"
    final_report_dir.mkdir(parents=True, exist_ok=True)

    for source_dir in sorted(p for p in base_download_dir.iterdir() if p.is_dir()):
        print(f"\n[INFO] Processing source directory: {source_dir.name}")

        material_dirs = sorted(
            p for p in source_dir.iterdir()
            if p.is_dir() and p.name.startswith("material_")
        )
        if not material_dirs:
            print(f"[INFO] No material folders found in {source_dir}")
            continue

        entries: List[Dict[str, Any]] = []
        for material_dir in material_dirs:
            entry = collect_material_entry(material_dir)
            if entry:
                entries.append(entry)

        summary_path = final_report_dir / f"{source_dir.name}{SUMMARY_SUFFIX}"
        write_summary_excel(entries, summary_path)


if __name__ == "__main__":
    generate_reports(BASE_DOWNLOAD_DIR)
